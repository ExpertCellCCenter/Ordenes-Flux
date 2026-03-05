import streamlit as st
import pandas as pd
import numpy as np
import unicodedata
import calendar
from datetime import date, datetime, time, timedelta
import pyodbc
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
from openpyxl.utils import get_column_letter

# -------------------------------------------------
# GLOBAL CONSTANTS
# -------------------------------------------------
EXCLUDED_VENDOR = "ABASTECEDORA Y SUMINISTROS ORTEGA/ISABEL VALDEZ JIMENEZ"
DEFAULT_LOOKBACK_DAYS = 30

FLOW_ORDER = [
    "Total",
    "Nuevo",
    "Back Office",
    "Solicitado",
    "En preparacion",
    "En entrega",
    "Reprogramado",
    "Entregado",
]
FLOW_STAGES_NO_TOTAL = ["Nuevo", "Back Office", "Solicitado", "En preparacion", "En entrega", "Reprogramado", "Entregado"]

# -------------------------------------------------
# STREAMLIT CONFIG
# -------------------------------------------------
st.set_page_config(
    page_title="Órdenes — Flujo & Tiempos",
    page_icon="⏱️",
    layout="wide",
)

# -------------------------------------------------
# STYLES
# -------------------------------------------------
st.markdown(
    """
<style>
html, body, [class*="css"] { font-family: "Segoe UI", system-ui, sans-serif; }
.block-container { padding-top: 1.1rem; padding-bottom: 2rem; }
h1 { font-weight: 900 !important; letter-spacing: -0.3px; }
h2, h3 { font-weight: 800 !important; }

/* --- Flow pills --- */
.flow-wrap {
  display:flex;
  gap: 10px;
  flex-wrap: wrap;
  align-items: center;
  padding: 10px 8px;
  border-radius: 16px;
  border: 1px solid rgba(148,163,184,0.35);
  background: rgba(148,163,184,0.06);
}
.flow-pill{
  min-width: 120px;
  text-align:center;
  padding: 10px 14px;
  border-radius: 12px;
  background: rgba(148,163,184,0.45);
  color: white;
  font-weight: 800;
  line-height: 1.05;
  box-shadow: 0 8px 16px rgba(15,23,42,0.10);
}
.flow-pill small{
  display:block;
  font-weight: 800;
  opacity: 0.95;
}
@media (prefers-color-scheme: dark) {
  .flow-wrap { background: rgba(15,23,42,0.55); border-color: rgba(148,163,184,0.45); }
  .flow-pill { background: rgba(148,163,184,0.35); box-shadow: 0 10px 22px rgba(0,0,0,0.45); }
}

/* KPI cards */
.kpi-row { display:flex; gap:12px; flex-wrap:wrap; margin-top: 10px; }
.kpi {
  background: rgba(255,255,255,0.92);
  border: 1px solid rgba(15,23,42,0.10);
  box-shadow: 0 10px 26px rgba(15,23,42,0.10);
  border-radius: 16px;
  padding: 14px 16px;
  min-width: 220px;
  flex: 1;
}
.kpi .label { font-size: 0.92rem; opacity:0.85; }
.kpi .value { font-size: 1.7rem; font-weight: 900; margin-top:6px; }
.kpi .sub { font-size: 0.85rem; opacity:0.75; margin-top:4px; }

@media (prefers-color-scheme: dark) {
  .kpi{
    background: rgba(15,23,42,0.92);
    border-color: rgba(148,163,184,0.55);
    box-shadow: 0 12px 34px rgba(0,0,0,0.55);
  }
}

/* Download button */
div[data-testid="stDownloadButton"] > button {
    border-radius: 999px;
    background: linear-gradient(90deg,#0ea5e9,#6366f1);
    color: #f9fafb;
    border: none;
    padding: 0.45rem 1.2rem;
    font-weight: 800;
}
div[data-testid="stDownloadButton"] > button:hover { filter: brightness(1.05); }

/* Plotly transparent */
.js-plotly-plot .plotly .main-svg { background-color: rgba(0,0,0,0) !important; }
</style>
""",
    unsafe_allow_html=True,
)

# -------------------------------------------------
# SESSION STATE
# -------------------------------------------------
if "last_refresh" not in st.session_state:
    st.session_state["last_refresh"] = datetime.now()

# -------------------------------------------------
# EXCEL EXPORT
# -------------------------------------------------
def _safe_sheet_name(name: str) -> str:
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    out = str(name)
    for ch in invalid:
        out = out.replace(ch, "_")
    out = out.strip() or "Sheet"
    return out[:31]


def dfs_to_excel_bytes(sheets: dict) -> bytes:
    output = BytesIO()
    used_names = set()

    def unique_sheet_name(base: str) -> str:
        base = _safe_sheet_name(base)
        if base not in used_names:
            used_names.add(base)
            return base
        i = 2
        while True:
            suffix = f"_{i}"
            cand = (base[: (31 - len(suffix))] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
            if cand not in used_names:
                used_names.add(cand)
                return cand
            i += 1

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for raw_name, df in sheets.items():
            sheet_name = unique_sheet_name(raw_name)
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column
            if max_col > 0 and max_row > 0:
                ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
                for col_idx in range(1, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    for cell in ws[col_letter]:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2

    output.seek(0)
    return output.getvalue()

# -------------------------------------------------
# DISPLAY HELPERS
# -------------------------------------------------
def fmt_int(x):
    try:
        return f"{int(x):,}"
    except Exception:
        return "0"


def fmt_pct(x):
    try:
        return f"{float(x) * 100:.1f}%"
    except Exception:
        return "0.0%"


def fmt_timedelta(td) -> str:
    if td is None or pd.isna(td):
        return "—"
    if isinstance(td, (float, int)):
        td = pd.to_timedelta(td, unit="s")
    if not isinstance(td, pd.Timedelta):
        try:
            td = pd.to_timedelta(td)
        except Exception:
            return "—"
    if td < pd.Timedelta(0):
        return "—"

    days = td.days
    secs = int(td.seconds)
    hh = secs // 3600
    mm = (secs % 3600) // 60
    return f"{days}d {hh:02d}:{mm:02d}" if days > 0 else f"{hh:02d}:{mm:02d}"


def fmt_done_or_in_process(td_done, td_age):
    if td_done is not None and pd.notna(td_done):
        return fmt_timedelta(td_done)
    if td_age is not None and pd.notna(td_age):
        return f"En proceso · {fmt_timedelta(td_age)}"
    return "—"


def td_to_hours(td) -> float:
    if td is None or pd.isna(td):
        return np.nan
    try:
        return float(pd.to_timedelta(td).total_seconds() / 3600.0)
    except Exception:
        return np.nan


def kpi_card(label, value, sub=None):
    sub_html = f'<div class="sub">{sub}</div>' if sub else ""
    st.markdown(
        f"""
        <div class="kpi">
          <div class="label">{label}</div>
          <div class="value">{value}</div>
          {sub_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_flow_pills(counts: dict):
    html = '<div class="flow-wrap">'
    for k in FLOW_ORDER:
        v = counts.get(k, 0)
        html += f'<div class="flow-pill">{k}<small>({int(v)})</small></div>'
        if k != FLOW_ORDER[-1]:
            html += '<div style="color: #94a3b8; font-weight: bold; font-size: 1.2rem;">➔</div>'
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)

# -------------------------------------------------
# TEXT NORMALIZATION & SANITIZATION (ANTI-1900)
# -------------------------------------------------
def _norm_col(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8", "ignore")
    s = s.replace("_", " ")
    s = " ".join(s.split())
    return s


def canon_estatus(x: str) -> str:
    s = str(x).strip()
    sn = _norm_col(s)
    if sn == "nuevo":
        return "Nuevo"
    if sn in ["back office", "backoffice"]:
        return "Back Office"
    if sn in ["solicitado", "solicitada"]:
        return "Solicitado"
    if sn in ["en preparacion", "en preparación", "en preparacion."]:
        return "En preparacion"
    if sn == "en entrega":
        return "En entrega"
    if sn in ["reprogramado", "reprogramada"]:
        return "Reprogramado"
    if sn in ["entregado", "entregada"]:
        return "Entregado"
    return str(x).strip()


def sanitize_dates(dt_series: pd.Series) -> pd.Series:
    if not pd.api.types.is_datetime64_any_dtype(dt_series):
        dt_series = pd.to_datetime(dt_series, errors="coerce")
    return dt_series.where(dt_series >= pd.Timestamp("2000-01-01"), pd.NaT)


def _extract_datetime_text(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    s = s.replace(
        {
            "nan": "", "none": "", "nat": "", "NaN": "", "None": "", "NaT": "", "<NA>": "", "null": "", "Null": "",
            "1900-01-01 00:00:00": "", "1900-01-01 00:00:00.000": "", "1900-01-01": "", "1900-01-01T00:00:00": "",
        }
    )
    s = s.where(s != "", np.nan)

    if s.notna().any():
        pat = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?)|(\d{4}[/-]\d{1,2}[/-]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?)"
        ext = s.astype(str).str.extract(pat)
        ext = ext[0].fillna(ext[1])
        s2 = ext.where(ext.notna(), s)
        return s2
    return s


def parse_dt_both(series: pd.Series) -> tuple:
    dt_native = pd.to_datetime(series, errors="coerce")
    s2 = _extract_datetime_text(series)
    dt_df = pd.to_datetime(s2, errors="coerce", dayfirst=True)
    dt_mf = pd.to_datetime(s2, errors="coerce", dayfirst=False)

    dt_df = dt_df.where(dt_df.notna(), dt_native)
    dt_mf = dt_mf.where(dt_mf.notna(), dt_native)
    return sanitize_dates(dt_df), sanitize_dates(dt_mf)

def choose_dt_rowwise(dt_df: pd.Series, dt_mf: pd.Series, created: pd.Series | None = None, bo: pd.Series | None = None) -> pd.Series:
    now_ts = pd.Timestamp(datetime.now())
    max_ts = now_ts + pd.Timedelta(days=1)

    out = dt_df.copy()
    
    valid_df = dt_df.notna() & (dt_df <= max_ts)
    valid_mf = dt_mf.notna() & (dt_mf <= max_ts)

    out = out.where(~(valid_mf & ~valid_df), dt_mf)

    base_dt = bo if bo is not None and bo.notna().any() else created
    
    if base_dt is not None:
        c_dt = pd.to_datetime(base_dt, errors="coerce")
        
        valid_df_c = valid_df & (c_dt.isna() | (dt_df >= c_dt))
        valid_mf_c = valid_mf & (c_dt.isna() | (dt_mf >= c_dt))
        
        out = out.where(~(valid_mf_c & ~valid_df_c), dt_mf)
        
        both = valid_df_c & valid_mf_c & c_dt.notna()
        if both.any():
            diff_df = (dt_df - c_dt).abs()
            diff_mf = (dt_mf - c_dt).abs()
            out = out.where(~(both & (diff_mf < diff_df)), dt_mf)

    out = out.where(out <= max_ts, pd.NaT)
    return sanitize_dates(out)


def choose_dt_activation_rowwise(dt_df: pd.Series, dt_mf: pd.Series, bo: pd.Series | None = None) -> pd.Series:
    now_ts = pd.Timestamp(datetime.now())
    max_ts = now_ts + pd.Timedelta(days=1)

    out = dt_df.copy()
    out = out.where(~(dt_df.isna() & dt_mf.notna()), dt_mf)

    valid_df = dt_df.notna() & (dt_df <= max_ts)
    valid_mf = dt_mf.notna() & (dt_mf <= max_ts)

    if bo is not None:
        bo_dt = pd.to_datetime(bo, errors="coerce")
        valid_df &= bo_dt.isna() | (dt_df >= bo_dt)
        valid_mf &= bo_dt.isna() | (dt_mf >= bo_dt)

        out = out.where(~(valid_mf & ~valid_df), dt_mf)

        both = valid_df & valid_mf & bo_dt.notna()
        if both.any():
            diff_df = (dt_df - bo_dt).abs()
            diff_mf = (dt_mf - bo_dt).abs()
            out = out.where(~(both & (diff_mf < diff_df)), dt_mf)
    else:
        out = out.where(~(valid_mf & ~valid_df), dt_mf)

    out = out.where(out <= max_ts, pd.NaT)
    return sanitize_dates(out)


def choose_dt_created_rowwise(
    dt_df: pd.Series,
    dt_mf: pd.Series,
    bo: pd.Series | None = None,
    window_start: date | None = None,
    window_end: date | None = None,
) -> pd.Series:
    now_ts = pd.Timestamp(datetime.now())
    max_ts = now_ts + pd.Timedelta(days=1)

    out = dt_df.copy()
    out = out.where(~(dt_df.isna() & dt_mf.notna()), dt_mf)

    valid_df = dt_df.notna() & (dt_df <= max_ts)
    valid_mf = dt_mf.notna() & (dt_mf <= max_ts)

    if bo is not None:
        bo_dt = pd.to_datetime(bo, errors="coerce")
        valid_df &= bo_dt.isna() | (dt_df <= bo_dt)
        valid_mf &= bo_dt.isna() | (dt_mf <= bo_dt)

        out = out.where(~(valid_mf & ~valid_df), dt_mf)

        both = valid_df & valid_mf & bo_dt.notna()
        if both.any():
            out = out.where(~(both & (dt_mf > dt_df)), dt_mf)
    else:
        if window_start is not None and window_end is not None:
            w0 = pd.Timestamp(window_start)
            w1 = pd.Timestamp(window_end) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)
            in_df = dt_df.between(w0, w1)
            in_mf = dt_mf.between(w0, w1)
            out = out.where(~(in_mf & ~in_df), dt_mf)

        out = out.where(~(valid_mf & ~valid_df), dt_mf)

    out = out.where(out <= max_ts, pd.NaT)
    return sanitize_dates(out)


def parse_backoffice_datetime(series: pd.Series, window_start: date | None = None, window_end: date | None = None) -> pd.Series:
    dt_df, dt_mf = parse_dt_both(series)

    if window_start is None or window_end is None:
        return dt_df

    w0 = pd.Timestamp(window_start)
    w1 = pd.Timestamp(window_end) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)

    in1 = dt_df.between(w0, w1)
    in2 = dt_mf.between(w0, w1)

    out = dt_df.copy()
    out = out.where(~(in2 & ~in1), dt_mf)
    out = out.where(~(dt_df.isna() & dt_mf.notna()), dt_mf)
    return out


def choose_backoffice_dt(df: pd.DataFrame, window_start: date, window_end: date) -> pd.Series:
    if "BO_DT_DF" in df.columns and "BO_DT_MF" in df.columns:
        dt_dayfirst = df["BO_DT_DF"]
        dt_monthfirst = df["BO_DT_MF"]

        w0 = pd.Timestamp(window_start)
        w1 = pd.Timestamp(window_end) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)

        in1 = dt_dayfirst.between(w0, w1)
        in2 = dt_monthfirst.between(w0, w1)

        out = dt_dayfirst.copy()
        out = out.where(~(in2 & ~in1), dt_monthfirst)
        out = out.where(~(dt_dayfirst.isna() & dt_monthfirst.notna()), dt_monthfirst)
        return sanitize_dates(out)

    return parse_backoffice_datetime(df["Back Office"], window_start=window_start, window_end=window_end)


def _reference_end_dt(fecha_fin: date) -> datetime:
    return datetime.now() if fecha_fin == date.today() else datetime.combine(fecha_fin, time(23, 59, 59))


def pick_activation_dt(df: pd.DataFrame) -> tuple:
    if df is None or df.empty:
        return pd.Series(pd.NaT, index=df.index), None

    cols_norm = {_norm_col(c): c for c in df.columns}
    bo = df["BO_DT"] if "BO_DT" in df.columns else None

    def _try_col(colname: str):
        if colname not in df.columns:
            return None
        dt_df, dt_mf = parse_dt_both(df[colname])
        chosen = choose_dt_activation_rowwise(dt_df, dt_mf, bo=bo)
        return chosen if chosen.notna().any() else None

    for key in ["fecha activacion", "fecha de activacion", "fecha activación"]:
        col = cols_norm.get(key)
        if col:
            chosen = _try_col(col)
            if chosen is not None:
                return chosen, col

    for key in ["fecha venta", "fecha de venta", "fecha_venta"]:
        col = cols_norm.get(key)
        if col:
            chosen = _try_col(col)
            if chosen is not None:
                return chosen, col

    col = cols_norm.get("venta")
    if col:
        chosen = _try_col(col)
        if chosen is not None:
            return chosen, col

    return pd.Series(pd.NaT, index=df.index), None


def pick_stage_dt_from_columns(df: pd.DataFrame, stage: str, created: pd.Series, bo: pd.Series) -> tuple:
    if df is None or df.empty:
        return pd.Series(pd.NaT, index=df.index), None

    stage_n = _norm_col(stage)

    kw = {
        "solicitado": ["solicitado", "solicit", "solic"],
        "en preparacion": ["preparacion", "preparación", "prep", "armado", "empaque"],
        "en entrega": ["en entrega", "entrega", "ruta", "transito", "trayecto"],
        "reprogramado": ["reprogramado", "reprog"],
        "entregado": ["entregado", "fecha entrega", "fecha entregado", "fin"],
    }
    keys = kw.get(stage_n, [stage_n])

    cols = []
    norm_cols = {c: _norm_col(c) for c in df.columns}

    for c, nc in norm_cols.items():
        if stage_n == "en entrega" and ("fecha entrega" in nc or "entregado" in nc):
            continue
        if stage_n == "entregado" and nc == "en entrega":
            continue
        if any(k in nc for k in keys):
            cols.append(c)

    best_col = None
    best_dt = pd.Series(pd.NaT, index=df.index)
    best_nonnull = -1

    for c in cols:
        dt_df, dt_mf = parse_dt_both(df[c])
        chosen = choose_dt_rowwise(dt_df, dt_mf, created=created, bo=bo)
        n = int(chosen.notna().sum())
        if n > best_nonnull:
            best_nonnull = n
            best_dt = chosen
            best_col = c

    return best_dt, best_col

# -------------------------------------------------
# FLOW COUNTS
# -------------------------------------------------
def compute_flow_counts(df: pd.DataFrame) -> dict:
    counts = {k: 0 for k in FLOW_ORDER}
    if df is None or df.empty or "Estatus" not in df.columns:
        return counts
    E = df["Estatus"].astype(str)
    counts["Total"] = int(len(df))
    for stg in FLOW_STAGES_NO_TOTAL:
        counts[stg] = int((E == stg).sum())
    return counts

# -------------------------------------------------
# DB CONNECTION
# -------------------------------------------------
@st.cache_resource
def get_connection():
    cfg = st.secrets["sql"]
    conn_str = (
        f"DRIVER={{{cfg['driver']}}};"
        f"SERVER={cfg['server']};"
        f"DATABASE={cfg['database']};"
        f"UID={cfg['user']};"
        f"PWD={cfg['password']};"
        "Encrypt=yes;"
        "TrustServerCertificate=yes;"
        "MARS_Connection=yes;"
    )
    return pyodbc.connect(conn_str, autocommit=True)

# -------------------------------------------------
# LOAD DATA FROM SQL
# -------------------------------------------------
@st.cache_data
def load_hoja1():
    sql = """
    SELECT DISTINCT
        e.[Nombre Completo] AS NombreCompleto,
        e.[Jefe Inmediato]  AS JefeDirecto,
        e.[Region],
        e.[SubRegion],
        e.[Plaza],
        e.[Tienda],
        e.[Puesto],
        e.[Canal de Venta],
        e.[Tipo Tienda],
        e.[Operacion],
        e.[Estatus]
    FROM reporte_empleado('EMPRESA_MAESTRA',1,'','') AS e
    WHERE
        e.[Canal de Venta] = 'ATT'
        AND e.[Operacion]   = 'CONTACT CENTER'
        AND e.[Tipo Tienda] = 'VIRTUAL'
        AND e.[Puesto] IN (
            'ASESOR TELEFONICO',
            'ASESOR TELEFONICO 7500',
            'EJECUTIVO TELEFONICO 6500 AM',
            'EJECUTIVO TELEFONICO 6500 PM',
            'SUPERVISOR DE CONTACT CENTER'
        )
        AND e.[Estatus] = 'ACTIVO';
    """
    conn = get_connection()
    df = pd.read_sql(sql, conn)

    text_cols = [
        "NombreCompleto", "JefeDirecto", "Region", "SubRegion", "Plaza", "Tienda",
        "Puesto", "Canal de Venta", "Tipo Tienda", "Operacion", "Estatus"
    ]
    for col in text_cols:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"nan": np.nan, "None": np.nan})

    df["JefeDirecto"] = df["JefeDirecto"].fillna("").str.strip().replace("", "ENCUBADORA")
    df["Coordinador"] = df["JefeDirecto"]
    df = df[df["NombreCompleto"].str.upper() != EXCLUDED_VENDOR].copy()
    return df


@st.cache_data
def load_consulta1(fecha_ini: date, fecha_fin: date) -> pd.DataFrame:
    fi = fecha_ini.strftime("%Y%m%d")
    ff = fecha_fin.strftime("%Y%m%d")

    sql = f"""
    SELECT
        *,
        [Tienda solicita] AS Centro
    FROM reporte_programacion_entrega('empresa_maestra', 4, '{fi}', '{ff}')
    WHERE
        [Tienda solicita] LIKE 'EXP ATT C CENTER%' AND
        [Estatus] IN ('Nuevo','Back Office','Solicitado','En preparacion','En entrega','Reprogramado','Entregado','Canc Error');
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SET NOCOUNT ON; SET ANSI_WARNINGS OFF;")
    df = pd.read_sql(sql, conn)
    cur.execute("SET ANSI_WARNINGS ON;")
    cur.close()
    return df


@st.cache_data
def load_rastreo_extra(fecha_ini: date, fecha_fin: date) -> pd.DataFrame:
    fi = fecha_ini.strftime("%Y%m%d")
    ff = fecha_fin.strftime("%Y%m%d")

    sql = f"""
    SELECT 
        r.id_pedido_telefonia AS Programacion, 
        r.accion, 
        MAX(r.fecha) AS fecha_rastreo
    FROM dbo.pedido_telefonia_rastreo r
    INNER JOIN reporte_programacion_entrega('empresa_maestra', 4, '{fi}', '{ff}') t
        ON r.id_pedido_telefonia = t.Programacion
    WHERE r.accion IN ('En preparacion', 'En entrega', 'Reprogramado')
    GROUP BY r.id_pedido_telefonia, r.accion
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SET NOCOUNT ON; SET ANSI_WARNINGS OFF;")
    df = pd.read_sql(sql, conn)
    cur.execute("SET ANSI_WARNINGS ON;")
    cur.close()
    return df

# -------------------------------------------------
# TRANSFORM
# -------------------------------------------------
def transform_consulta1(df_raw: pd.DataFrame, hoja: pd.DataFrame, rastreo_extra: pd.DataFrame) -> pd.DataFrame:
    df = df_raw.copy()

    if rastreo_extra is not None and not rastreo_extra.empty:
        piv = rastreo_extra.pivot_table(index="Programacion", columns="accion", values="fecha_rastreo", aggfunc="max").reset_index()
        rename_map = {
            "En preparacion": "Fecha En preparacion Exacta",
            "En entrega": "Fecha En entrega Exacta",
            "Reprogramado": "Fecha Reprogramado Exacta",
        }
        piv.rename(columns=rename_map, inplace=True)
        if "Programacion" in df.columns:
            df = df.merge(piv, on="Programacion", how="left")

    clean_cols = [
        "Centro", "Estatus", "Back Office", "Vendedor", "Cliente",
        "Nuevo", "Solicitado", "En preparacion", "En preparación",
        "En entrega", "Reprogramado", "Entregado", "Fecha creacion", "Venta",
    ]
    for col in df.columns:
        if col in clean_cols or "fecha" in col.lower() or any(stg in col for stg in ["Solicitado", "preparacion", "entrega", "Reprogramado", "Entregado"]):
            df[col] = df[col].astype(str).str.strip().replace(
                {
                    "nan": np.nan, "None": np.nan, "NaT": np.nan, "nat": np.nan, "none": np.nan, "<NA>": np.nan, "null": np.nan,
                    "1900-01-01 00:00:00": np.nan, "1900-01-01 00:00:00.000": np.nan, "1900-01-01": np.nan, "1900-01-01T00:00:00": np.nan,
                }
            )

    if "Vendedor" in df.columns:
        df = df[df["Vendedor"].astype(str).str.upper() != EXCLUDED_VENDOR].copy()

    if "Estatus" in df.columns:
        df["Estatus"] = df["Estatus"].astype(str).map(canon_estatus)

    df["Centro Original"] = pd.Series(pd.NA, index=df.index, dtype="object")
    mask_cc2 = df["Centro"].astype(str).str.contains("EXP ATT C CENTER 2", na=False)
    mask_jv = df["Centro"].astype(str).str.contains("EXP ATT C CENTER JUAREZ", na=False)
    df.loc[mask_cc2, "Centro Original"] = "CC2"
    df.loc[mask_jv, "Centro Original"] = "CC JV"

    empleados_join = hoja[hoja["Puesto"].isin(["ASESOR TELEFONICO 7500", "EJECUTIVO TELEFONICO 6500 AM"])].copy()
    empleados_join = empleados_join[empleados_join["JefeDirecto"] != "ENCUBADORA"].drop_duplicates(subset=["NombreCompleto"])
    empleados_join = empleados_join[empleados_join["NombreCompleto"].str.upper() != EXCLUDED_VENDOR]

    hoja_join = empleados_join.rename(columns={"NombreCompleto": "Nombre Completo", "JefeDirecto": "Jefe directo"})
    df = df.merge(
        hoja_join[["Nombre Completo", "Jefe directo", "Coordinador"]],
        how="left",
        left_on="Vendedor",
        right_on="Nombre Completo",
    )
    df.drop(columns=["Nombre Completo"], inplace=True, errors="ignore")
    df["Jefe directo"] = df["Jefe directo"].fillna("").astype(str).str.strip().replace("", "ENCUBADORA")

    if "Fecha creacion" in df.columns:
        df["Fecha creacion"] = pd.to_datetime(df["Fecha creacion"], errors="coerce", dayfirst=True)
        df["Fecha creacion"] = sanitize_dates(df["Fecha creacion"])

    if "Back Office" in df.columns:
        s = df["Back Office"].astype(str).str.strip()
        s = s.replace({"nan": "", "None": "", "NaT": "", "<NA>": "", "null": ""})
        s = s.where(s != "", np.nan)
        if s.notna().any():
            pat = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s+\d{1,2}:\d{2}(?::\d{2})?)|(\d{4}[/-]\d{1,2}[/-]\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?)"
            ext = s.astype(str).str.extract(pat)
            ext = ext[0].fillna(ext[1])
            s2 = ext.where(ext.notna(), s)
        else:
            s2 = s

        df["BO_DT_DF"] = pd.to_datetime(s2, errors="coerce", dayfirst=True)
        df["BO_DT_DF"] = sanitize_dates(df["BO_DT_DF"])

        df["BO_DT_MF"] = pd.to_datetime(s2, errors="coerce", dayfirst=False)
        df["BO_DT_MF"] = sanitize_dates(df["BO_DT_MF"])

    return df

# -------------------------------------------------
# BUILD VIEW
# -------------------------------------------------
def build_view(df_ctx: pd.DataFrame, fecha_ini: date, fecha_fin: date):
    meta = {"activation_col": None, "has_activation_dt": False, "stage_sources": {}}
    df = df_ctx.copy()

    df["BO_DT"] = choose_backoffice_dt(df, window_start=fecha_ini, window_end=fecha_fin) if "Back Office" in df.columns else pd.NaT

    if "Fecha creacion" in df.columns:
        c_df, c_mf = parse_dt_both(df["Fecha creacion"])
        df["CREATED_DT"] = choose_dt_created_rowwise(c_df, c_mf, bo=df["BO_DT"], window_start=fecha_ini, window_end=fecha_fin)
    else:
        df["CREATED_DT"] = pd.NaT

    meta["stage_sources"]["Nuevo"] = "Fecha creacion" if "Fecha creacion" in df.columns else None
    meta["stage_sources"]["Back Office"] = "Back Office" if "Back Office" in df.columns else None

    df["STG_Nuevo_DT"] = df["CREATED_DT"]
    df["STG_BackOffice_DT"] = df["BO_DT"]

    for stage, outcol in [
        ("Solicitado", "STG_Solicitado_DT"),
        ("En preparacion", "STG_EnPreparacion_DT"),
        ("En entrega", "STG_EnEntrega_DT"),
        ("Reprogramado", "STG_Reprogramado_DT"),
        ("Entregado", "STG_Entregado_DT"),
    ]:
        dt_stage, src = pick_stage_dt_from_columns(df, stage, created=df["CREATED_DT"], bo=df["BO_DT"])
        df[outcol] = sanitize_dates(dt_stage)
        meta["stage_sources"][stage] = src

    act_dt, act_col = pick_activation_dt(df)
    df["ACT_DT"] = sanitize_dates(act_dt)

    _now = pd.Timestamp(datetime.now()) + pd.Timedelta(days=1)
    df.loc[df["ACT_DT"] > _now, "ACT_DT"] = pd.NaT

    meta["activation_col"] = act_col
    meta["has_activation_dt"] = bool(df["ACT_DT"].notna().any())

    if "Venta" in df.columns:
        venta = df["Venta"]
        venta_ok = ~(venta.isna() | venta.astype(str).str.strip().eq(""))
    else:
        venta_ok = pd.Series(False, index=df.index)

    df["IS_ACTIVADA_COMPLETA"] = venta_ok | df["ACT_DT"].notna()
    df["ENTREGADO_SIN_VENTA"] = (df["Estatus"].astype(str).eq("Entregado")) & (~venta_ok)

    df["TD_Creacion_a_BO"] = df["BO_DT"] - df["CREATED_DT"]
    df["TD_BO_a_Act"] = df["ACT_DT"] - df["BO_DT"]
    df["TD_Creacion_a_Act"] = df["ACT_DT"] - df["CREATED_DT"]

    def safe_td(a, b):
        td = a - b
        td = td.where(td >= pd.Timedelta(0))
        return td

    df["TD_BO_a_Solicitado"] = safe_td(df["STG_Solicitado_DT"], df["STG_BackOffice_DT"])
    df["TD_Solicitado_a_Preparacion"] = safe_td(df["STG_EnPreparacion_DT"], df["STG_Solicitado_DT"])
    df["TD_Preparacion_a_EnEntrega"] = safe_td(df["STG_EnEntrega_DT"], df["STG_EnPreparacion_DT"])
    df["TD_EnEntrega_a_Entregado"] = safe_td(df["STG_Entregado_DT"], df["STG_EnEntrega_DT"])

    df["TD_BO_a_Entregado"] = safe_td(df["STG_Entregado_DT"], df["BO_DT"])
    df["TD_Creacion_a_Entregado"] = safe_td(df["STG_Entregado_DT"], df["CREATED_DT"])

    ref_dt = pd.Timestamp(_reference_end_dt(fecha_fin))
    df["TD_Age_Desde_Creacion"] = ref_dt - df["CREATED_DT"]
    df["TD_Age_Desde_BO"] = ref_dt - df["BO_DT"]

    for c in [
        "TD_Creacion_a_BO", "TD_BO_a_Act", "TD_Creacion_a_Act",
        "TD_Age_Desde_Creacion", "TD_Age_Desde_BO",
        "TD_BO_a_Entregado", "TD_Creacion_a_Entregado",
    ]:
        df.loc[df[c] < pd.Timedelta(0), c] = pd.NaT

    df["H_Creacion_a_BO"] = df["TD_Creacion_a_BO"].apply(td_to_hours)
    df["H_BO_a_Act"] = df["TD_BO_a_Act"].apply(td_to_hours)
    df["H_Creacion_a_Act"] = df["TD_Creacion_a_Act"].apply(td_to_hours)
    df["H_Age_Desde_Creacion"] = df["TD_Age_Desde_Creacion"].apply(td_to_hours)
    df["H_Age_Desde_BO"] = df["TD_Age_Desde_BO"].apply(td_to_hours)
    
    df["H_Nuevo_a_BO"] = df["TD_Creacion_a_BO"].apply(td_to_hours)
    df["H_BO_a_Solicitado"] = df["TD_BO_a_Solicitado"].apply(td_to_hours)
    df["H_Solicitado_a_Prep"] = df["TD_Solicitado_a_Preparacion"].apply(td_to_hours)
    df["H_Prep_a_Entrega"] = df["TD_Preparacion_a_EnEntrega"].apply(td_to_hours)
    df["H_Entrega_a_Entregado"] = df["TD_EnEntrega_a_Entregado"].apply(td_to_hours)

    df["CREATED_DATE"] = df["CREATED_DT"].dt.date
    df["CREATED_HOUR"] = df["CREATED_DT"].dt.hour
    df["CREATED_DOW"] = df["CREATED_DT"].dt.day_name()

    return df, meta

# -------------------------------------------------
# VISUALS: buckets + top slow + scatter
# -------------------------------------------------
def _bucket_hours(h: float) -> str:
    if h is None or (isinstance(h, float) and np.isnan(h)):
        return "Sin dato"
    if h <= 2:
        return "≤2h"
    if h <= 6:
        return "2–6h"
    if h <= 12:
        return "6–12h"
    if h <= 24:
        return "12–24h"
    if h <= 48:
        return "1–2d"
    if h <= 72:
        return "2–3d"
    return ">3d"


_BUCKET_ORDER = ["≤2h", "2–6h", "6–12h", "12–24h", "1–2d", "2–3d", ">3d", "Sin dato"]


def make_time_buckets_chart(view: pd.DataFrame) -> go.Figure | None:
    if view.empty:
        return None

    d = view.copy()
    d_ent = d[d["Estatus"].astype(str).eq("Entregado")].copy()

    if d_ent.empty and d["TD_Creacion_a_BO"].notna().sum() == 0:
        return None

    rows = []

    if "TD_Creacion_a_BO" in d.columns:
        h = d["TD_Creacion_a_BO"].apply(td_to_hours)
        vc = h.apply(_bucket_hours).value_counts().to_dict()
        for k, v in vc.items():
            rows.append({"Tramo": "Nuevo→BO", "Rango": k, "Órdenes": int(v)})

    if not d_ent.empty and "TD_BO_a_Entregado" in d_ent.columns:
        h = d_ent["TD_BO_a_Entregado"].apply(td_to_hours)
        vc = h.apply(_bucket_hours).value_counts().to_dict()
        for k, v in vc.items():
            rows.append({"Tramo": "BO→Entregado", "Rango": k, "Órdenes": int(v)})

    if not d_ent.empty and "TD_Creacion_a_Entregado" in d_ent.columns:
        h = d_ent["TD_Creacion_a_Entregado"].apply(td_to_hours)
        vc = h.apply(_bucket_hours).value_counts().to_dict()
        for k, v in vc.items():
            rows.append({"Tramo": "Total→Entregado", "Rango": k, "Órdenes": int(v)})

    if not rows:
        return None

    dfb = pd.DataFrame(rows)
    dfb["Rango"] = pd.Categorical(dfb["Rango"], categories=_BUCKET_ORDER, ordered=True)
    dfb = dfb.sort_values(["Rango", "Tramo"])

    fig = px.bar(
        dfb,
        x="Rango",
        y="Órdenes",
        color="Tramo",
        barmode="group",
        text_auto=True, 
        title="Distribución por rangos de tiempo (volumen general)",
        template="plotly_white",
        color_discrete_sequence=["#0ea5e9", "#6366f1", "#10b981"]
    )
    
    fig.update_traces(textfont_size=13, textangle=0, textposition="outside", cliponaxis=False)
    fig.update_layout(
        margin=dict(l=40, r=20, t=70, b=20), 
        legend_title_text="Tramo",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    fig.update_yaxes(title="Cantidad de Órdenes", showgrid=True, gridcolor="#f1f5f9")
    fig.update_xaxes(title="Rango de Tiempo", showgrid=False)
    
    return fig


def make_bottleneck_matrix(view: pd.DataFrame) -> go.Figure | None:
    if view.empty:
        return None

    stages_map = {
        "1. Nuevo→BO": "H_Nuevo_a_BO",
        "2. BO→Solicit": "H_BO_a_Solicitado",
        "3. Solicit→Prep": "H_Solicitado_a_Prep",
        "4. Prep→Entrega": "H_Prep_a_Entrega",
        "5. Entrega→Fin": "H_Entrega_a_Entregado"
    }

    rows = []
    for stage_name, col in stages_map.items():
        if col in view.columns:
            valid_hours = view[view[col].notna()][col]
            if not valid_hours.empty:
                buckets = valid_hours.apply(_bucket_hours).value_counts(normalize=True) * 100
                for bucket, pct in buckets.items():
                    rows.append({"Etapa": stage_name, "Rango": bucket, "Porcentaje": pct})

    if not rows:
        return None

    df_b = pd.DataFrame(rows)
    df_b["Rango"] = pd.Categorical(df_b["Rango"], categories=_BUCKET_ORDER, ordered=True)
    df_b = df_b.sort_values(["Etapa", "Rango"])

    color_map = {
        "≤2h": "#10b981",     
        "2–6h": "#34d399",    
        "6–12h": "#fcd34d",   
        "12–24h": "#fbbf24",  
        "1–2d": "#f97316",    
        "2–3d": "#ef4444",    
        ">3d": "#991b1b",     
        "Sin dato": "#e2e8f0"
    }

    fig = px.bar(
        df_b,
        x="Etapa",
        y="Porcentaje",
        color="Rango",
        title="🚦 Salud del Proceso: ¿Dónde se atoran las órdenes? (% por etapa)",
        color_discrete_map=color_map,
        text_auto=".1f" 
    )

    fig.update_traces(textfont_size=12, textfont_color="white")
    fig.update_layout(
        template="plotly_white",
        margin=dict(l=40, r=20, t=70, b=20),
        yaxis_title="% de Órdenes",
        xaxis_title="Transición de Etapa",
        legend_title="Tiempo que tardó"
    )
    return fig


def _pick_order_id_col(df: pd.DataFrame) -> str:
    for c in ["Programacion", "Folio", "Telefono", "Cliente"]:
        if c in df.columns:
            return c
    return "__index__"


def make_top_slowest_bar(view: pd.DataFrame, n: int = 20) -> go.Figure | None:
    if view.empty or "TD_Creacion_a_Entregado" not in view.columns:
        return None
    d = view[view["Estatus"].astype(str).eq("Entregado")].copy()
    if d.empty:
        return None

    d["_h_total"] = d["TD_Creacion_a_Entregado"].apply(td_to_hours)
    d = d[np.isfinite(d["_h_total"])].copy()
    if d.empty:
        return None

    id_col = _pick_order_id_col(d)
    if id_col == "__index__":
        d[id_col] = d.index.astype(str)
    else:
        d[id_col] = d[id_col].astype(str)

    d = d.sort_values("_h_total", ascending=False).head(int(n)).copy()
    d = d.sort_values("_h_total", ascending=True)

    stage_cols = {
        "H_Nuevo_a_BO": "1. Nuevo→BO",
        "H_BO_a_Solicitado": "2. BO→Solicit",
        "H_Solicitado_a_Prep": "3. Solicit→Prep",
        "H_Prep_a_Entrega": "4. Prep→Entrega",
        "H_Entrega_a_Entregado": "5. Entrega→Fin"
    }
    
    melted = []
    for _, row in d.iterrows():
        order_id = row[id_col]
        for col, label in stage_cols.items():
            if col in row and pd.notna(row[col]):
                val = float(row[col])
                if val > 0: 
                    melted.append({
                        "Orden": order_id,
                        "Etapa": label,
                        "Horas": val,
                        "Total_Horas": row["_h_total"] 
                    })
                    
    df_melt = pd.DataFrame(melted)
    if df_melt.empty:
        return None

    fig = px.bar(
        df_melt,
        x="Horas",
        y="Orden",
        color="Etapa",
        orientation="h",
        title=f"🚨 Radiografía: Top {len(d)} Órdenes Más Lentas (Desglose por Etapa)",
        template="plotly_white",
        text_auto=".1f",
        color_discrete_sequence=px.colors.qualitative.Pastel
    )
    
    fig.update_traces(textfont_size=11, textposition="inside")
    fig.update_layout(
        margin=dict(l=40, r=40, t=70, b=20), 
        barmode='stack',
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    fig.update_xaxes(title="Horas Acumuladas (Creación → Entregado)", showgrid=True, gridcolor="#f1f5f9")
    fig.update_yaxes(title="ID de Orden")
    
    return fig


def make_scatter_orders(view: pd.DataFrame, color_by: str | None = None) -> go.Figure | None:
    if view.empty:
        return None
    d = view[view["Estatus"].astype(str).eq("Entregado")].copy()
    if d.empty or "STG_Entregado_DT" not in d.columns:
        return None
    d = d[d["STG_Entregado_DT"].notna()].copy()
    if d.empty:
        return None

    d["_h_total"] = d["TD_Creacion_a_Entregado"].apply(td_to_hours)
    d = d[np.isfinite(d["_h_total"])].copy()
    if d.empty:
        return None

    d = d.sort_values("STG_Entregado_DT", ascending=True).tail(2500).copy()

    id_col = _pick_order_id_col(d)
    if id_col == "__index__":
        d[id_col] = d.index.astype(str)
    else:
        d[id_col] = d[id_col].astype(str)

    hover_cols = [c for c in [id_col, "Cliente", "Telefono", "Folio", "Programacion", "Centro Original", "Jefe directo", "Vendedor"] if c in d.columns]
    hover = {c: True for c in hover_cols}
    hover["_h_total"] = True

    if color_by is None or color_by not in d.columns:
        fig = px.scatter(
            d,
            x="STG_Entregado_DT",
            y="_h_total",
            title="Cada orden: Fecha Entregado vs Horas Totales (Creación→Entregado)",
            hover_data=hover,
        )
    else:
        fig = px.scatter(
            d,
            x="STG_Entregado_DT",
            y="_h_total",
            color=color_by,
            title=f"Cada orden: Fecha Entregado vs Horas Totales (color por {color_by})",
            hover_data=hover,
        )

    fig.update_traces(
        marker=dict(size=9, opacity=0.6, line=dict(width=1, color='rgba(0,0,0,0.2)'))
    )
    fig.update_layout(
        template="plotly_white",
        margin=dict(l=40, r=20, t=70, b=20),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    fig.update_xaxes(title="Fecha de Entrega", showgrid=True, gridcolor="#f1f5f9")
    fig.update_yaxes(title="Horas (Creación → Entregado)", showgrid=True, gridcolor="#f1f5f9", zeroline=False)
    return fig

# -------------------------------------------------
# OTHER CHARTS
# -------------------------------------------------
def make_funnel(counts: dict) -> go.Figure:
    stages = ["Nuevo", "Back Office", "Solicitado", "En preparacion", "En entrega", "Reprogramado", "Entregado"]
    values = [int(counts.get(s, 0)) for s in stages]
    fig = go.Figure(go.Funnel(y=stages, x=values))
    fig.update_layout(title="Funnel del flujo operativo (conteos por etapa)", margin=dict(l=40, r=20, t=60, b=20), template="plotly_white")
    return fig


def make_flow_bar(counts: dict) -> go.Figure:
    stages = ["Nuevo", "Back Office", "Solicitado", "En preparacion", "En entrega", "Reprogramado", "Entregado"]
    values = [int(counts.get(s, 0)) for s in stages]
    fig = px.bar(pd.DataFrame({"Etapa": stages, "Total": values}), x="Etapa", y="Total", title="Conteos por etapa", template="plotly_white", text_auto=True)
    fig.update_xaxes(type="category")
    fig.update_layout(margin=dict(l=40, r=20, t=60, b=20))
    return fig


def make_backlog_over_time(view: pd.DataFrame) -> go.Figure | None:
    if view.empty or view["CREATED_DT"].isna().all():
        return None
    dfp = view.copy()
    dfp["Fecha"] = dfp["CREATED_DT"].dt.date
    grp = dfp.groupby(["Fecha", "Estatus"]).size().reset_index(name="Total")
    grp = grp[grp["Estatus"].isin(FLOW_STAGES_NO_TOTAL)].copy()
    fig = px.area(grp, x="Fecha", y="Total", color="Estatus", title="Backlog por etapa a través del tiempo (creación)", template="plotly_white")
    return fig


def make_heatmap_created(view: pd.DataFrame) -> go.Figure | None:
    if view.empty or view["CREATED_DT"].isna().all():
        return None
    tmp = view.copy()
    tmp["DOW"] = tmp["CREATED_DT"].dt.day_name()
    tmp["HOUR"] = tmp["CREATED_DT"].dt.hour
    piv = tmp.pivot_table(index="DOW", columns="HOUR", values="Estatus", aggfunc="count", fill_value=0)
    order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    piv = piv.reindex([d for d in order if d in piv.index])
    fig = px.imshow(piv, title="Mapa de calor: órdenes creadas (día vs hora)", aspect="auto")
    fig.update_layout(margin=dict(l=40, r=20, t=60, b=20))
    return fig

# -------------------------------------------------
# MAIN
# -------------------------------------------------
def _month_bounds(y: int, m: int) -> tuple[date, date]:
    last_day = calendar.monthrange(y, m)[1]
    return date(y, m, 1), date(y, m, last_day)


def main():
    st.title("⏱️ Órdenes — Flujo y Tiempo (Nuevo → BO → Entregado)")

    st.sidebar.header("Panel de control")

    if st.sidebar.button("🔄 Actualizar"):
        st.cache_data.clear()
        st.cache_resource.clear()
        st.session_state["last_refresh"] = datetime.now()
        st.rerun()

    # ✅ NUEVO: FILTRO DE PERIODO "DUMMY-FRIENDLY"
    st.sidebar.markdown("---")
    st.sidebar.subheader("📅 Periodo de Análisis")
    
    today = date.today()
    years = list(range(today.year - 2, today.year + 1))
    meses_nombres = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]

    c1, c2 = st.sidebar.columns(2)
    with c1:
        y_sel = st.selectbox("Año", years, index=years.index(today.year))
    with c2:
        m_sel = st.selectbox("Mes", meses_nombres, index=today.month - 1)

    # Convertir el nombre del mes a número (1-12)
    m_num = meses_nombres.index(m_sel) + 1

    # Calcular las fechas automáticamente para la base de datos
    m0, m1 = _month_bounds(int(y_sel), m_num)
    fecha_ini = m0
    
    # Si seleccionan el mes y año en curso, el límite es hoy para no buscar en el futuro
    if int(y_sel) == today.year and m_num == today.month:
        fecha_fin = min(m1, today)
    else:
        fecha_fin = m1

    if fecha_ini > today:
        st.sidebar.warning("Estás seleccionando un mes en el futuro.")

    with st.spinner("Cargando datos..."):
        hoja = load_hoja1()
        raw = load_consulta1(fecha_ini, fecha_fin)
        rastreo_extra = load_rastreo_extra(fecha_ini, fecha_fin)
        consulta = transform_consulta1(raw, hoja, rastreo_extra)

    # ✅ NUEVO: FILTROS LIMPIOS Y EN CASCADA (FUERA DEL EXPANDER)
    st.sidebar.markdown("---")
    st.sidebar.subheader("🔎 Filtros de Operación")
    
    df_filter = consulta.copy()

    # 1. Filtro: Centro
    if "Centro Original" in df_filter.columns:
        centros = ["Todos"] + sorted([str(c) for c in df_filter["Centro Original"].dropna().unique() if str(c).strip() != ""])
        centro_sel = st.sidebar.selectbox("🏢 Centro", centros, index=0)
        if centro_sel != "Todos":
            df_filter = df_filter[df_filter["Centro Original"] == centro_sel]
    else:
        centro_sel = "Todos"

    # 2. Filtro: Supervisor (Depende del Centro seleccionado)
    if "Jefe directo" in df_filter.columns:
        supervisores = ["Todos"] + sorted([str(s) for s in df_filter["Jefe directo"].dropna().unique() if str(s).strip() != ""])
        supervisor_sel = st.sidebar.selectbox("👤 Supervisor", supervisores, index=0)
        if supervisor_sel != "Todos":
            df_filter = df_filter[df_filter["Jefe directo"] == supervisor_sel]
    else:
        supervisor_sel = "Todos"

    # 3. Filtro: Ejecutivo (Depende del Centro y Supervisor seleccionados)
    if "Vendedor" in df_filter.columns:
        ejecutivos = ["Todos"] + sorted([str(v) for v in df_filter["Vendedor"].dropna().unique() if str(v).strip() != ""])
        ejecutivo_sel = st.sidebar.selectbox("🎧 Ejecutivo", ejecutivos, index=0)
        if ejecutivo_sel != "Todos":
            df_filter = df_filter[df_filter["Vendedor"] == ejecutivo_sel]
    else:
        ejecutivo_sel = "Todos"

    # El DataFrame final que se usa en toda la app
    df = df_filter.copy()

    # ✅ OCULTO: Las alertas ahora están en un expander para no ensuciar el menú principal
    st.sidebar.markdown("---")
    with st.sidebar.expander("⚙️ Configuración de Alertas (Horas)"):
        st.caption("Define el umbral para considerar una orden como 'Crítica'.")
        alert_map = {
            "Nuevo": st.number_input("Nuevo >", 1, 720, 24, 1),
            "Back Office": st.number_input("Back Office >", 1, 720, 24, 1),
            "Solicitado": st.number_input("Solicitado >", 1, 720, 24, 1),
            "En preparacion": st.number_input("En preparacion >", 1, 720, 24, 1),
            "En entrega": st.number_input("En entrega >", 1, 720, 48, 1),
            "Reprogramado": st.number_input("Reprogramado >", 1, 720, 48, 1),
        }

    tabs = st.tabs(["Resumen Ejecutivo", "Gráficas", "Pendientes a Recuperar", "Detalle / Export"])

    # ============================
    # TAB 0: RESUMEN (visual + per-order)
    # ============================
    with tabs[0]:
        if df.empty:
            st.info("No hay datos para los filtros seleccionados.")
        else:
            view, meta = build_view(df, fecha_ini, fecha_fin)
            counts = compute_flow_counts(view)

            st.subheader("Flujo de Órdenes (operación)")
            render_flow_pills(counts)

            total = int(len(view))
            entregado = int((view["Estatus"].astype(str).eq("Entregado")).sum())
            pendientes = int((~view["Estatus"].astype(str).eq("Entregado")).sum())
            entregado_sin_venta_cnt = int(view["ENTREGADO_SIN_VENTA"].sum())

            st.markdown('<div class="kpi-row">', unsafe_allow_html=True)
            kpi_card("Órdenes en el periodo", fmt_int(total), sub=f"Del {fecha_ini} al {fecha_fin}")
            kpi_card("Entregadas", fmt_int(entregado))
            kpi_card("Pendientes", fmt_int(pendientes))
            kpi_card("Entregado sin Venta", fmt_int(entregado_sin_venta_cnt))
            st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("---")
            st.subheader("🎯 Tiempos del Proceso General")

            c1, c2 = st.columns(2)
            with c1:
                fig_buckets = make_time_buckets_chart(view)
                if fig_buckets is not None:
                    st.plotly_chart(fig_buckets, use_container_width=True, key="t0_buckets")
                else:
                    st.info("No hay datos suficientes para distribución por rangos.")

            with c2:
                fig_health = make_bottleneck_matrix(view)
                if fig_health is not None:
                    st.plotly_chart(fig_health, use_container_width=True, key="t0_health")
                else:
                    st.info("No hay datos suficientes para la matriz de cuellos de botella.")

            st.markdown("---")
            st.subheader("🕵️ Análisis de Lentas y Cuellos de Botella")

            fig_top = make_top_slowest_bar(view, n=20)
            if fig_top is not None:
                st.plotly_chart(fig_top, use_container_width=True, key="t0_top_slowest")
            else:
                st.info("No hay entregadas con tiempo total para ranking.")

            st.markdown("---")
            st.subheader("🟣 Cada orden (visual): fecha entregado vs horas totales")

            color_opts = ["(sin color)"]
            for c in ["Centro Original", "Jefe directo", "Vendedor"]:
                if c in view.columns:
                    color_opts.append(c)
            color_by = st.selectbox("Color por", color_opts, index=1 if len(color_opts) > 1 else 0)
            color_by = None if color_by == "(sin color)" else color_by

            fig_sc = make_scatter_orders(view, color_by=color_by)
            if fig_sc is not None:
                st.plotly_chart(fig_sc, use_container_width=True, key="t0_scatter")
            else:
                st.info("No hay suficientes entregadas con fecha entregado para el scatter.")

            st.markdown("---")
            st.subheader("📄 Tiempos por orden (Nuevo → BO → Entregado)")
            st.caption("Aquí ves **cada orden** y cuánto tardó en cada tramo (o 'En proceso').")

            show_mode = st.selectbox("Mostrar", ["Solo Entregadas", "Solo Pendientes", "Todas"], index=0)
            sort_mode = st.selectbox(
                "Ordenar por",
                ["Más recientes", "Más lentas Nuevo→BO", "Más lentas BO→Entregado", "Más lentas Total a Entregado"],
                index=0,
            )
            topn = st.slider("Cantidad a mostrar", 20, 300, 80, 10)

            tt = view.copy()

            tt["Tiempo Nuevo→BO"] = [
                (fmt_timedelta(done) if (done is not None and pd.notna(done))
                 else (f"En proceso · {fmt_timedelta(age)}" if (str(stg) == "Nuevo" and age is not None and pd.notna(age)) else "—"))
                for done, age, stg in zip(tt["TD_Creacion_a_BO"], tt["TD_Age_Desde_Creacion"], tt["Estatus"])
            ]
            tt["Tiempo BO→Entregado"] = [
                fmt_done_or_in_process(done, age)
                for done, age in zip(tt["TD_BO_a_Entregado"], tt["TD_Age_Desde_BO"])
            ]
            tt["Tiempo Total a Entregado"] = [
                fmt_done_or_in_process(done, age)
                for done, age in zip(tt["TD_Creacion_a_Entregado"], tt["TD_Age_Desde_Creacion"])
            ]

            if show_mode == "Solo Entregadas":
                tt = tt[tt["Estatus"].astype(str).eq("Entregado")].copy()
            elif show_mode == "Solo Pendientes":
                tt = tt[~tt["Estatus"].astype(str).eq("Entregado")].copy()

            tt["_h_nb"] = tt["TD_Creacion_a_BO"].apply(td_to_hours)
            tt["_h_be"] = tt["TD_BO_a_Entregado"].apply(td_to_hours)
            tt["_h_te"] = tt["TD_Creacion_a_Entregado"].apply(td_to_hours)

            if sort_mode == "Más lentas Nuevo→BO":
                tt = tt.sort_values("_h_nb", ascending=False)
            elif sort_mode == "Más lentas BO→Entregado":
                tt = tt.sort_values("_h_be", ascending=False)
            elif sort_mode == "Más lentas Total a Entregado":
                tt = tt.sort_values("_h_te", ascending=False)
            else:
                if "STG_Entregado_DT" in tt.columns and tt["STG_Entregado_DT"].notna().any():
                    tt = tt.sort_values("STG_Entregado_DT", ascending=False)
                else:
                    tt = tt.sort_values("CREATED_DT", ascending=False)

            cols_main = [c for c in [
                "Estatus",
                "Centro Original", "Jefe directo", "Vendedor",
                "Cliente", "Telefono", "Folio", "Programacion",
                "CREATED_DT", "BO_DT", "STG_Entregado_DT",
                "Tiempo Nuevo→BO", "Tiempo BO→Entregado", "Tiempo Total a Entregado",
                "Venta", "ENTREGADO_SIN_VENTA",
            ] if c in tt.columns]

            show = tt[cols_main].head(int(topn)).copy().rename(
                columns={
                    "Vendedor": "Ejecutivo",
                    "CREATED_DT": "Fecha Creación",
                    "BO_DT": "Fecha Back Office",
                    "STG_Entregado_DT": "Fecha Entregado",
                }
            )
            st.dataframe(show, use_container_width=True, height=520, hide_index=True)

            if entregado_sin_venta_cnt > 0:
                st.warning(f"⚠️ Hay **{entregado_sin_venta_cnt}** órdenes en **Entregado** pero **sin Venta** (revisar / recuperar).")
                with st.expander("Ver cuáles son Entregado sin Venta", expanded=True):
                    df_esv = view[view["ENTREGADO_SIN_VENTA"]].copy()
                    df_esv["Antigüedad"] = df_esv["TD_Age_Desde_Creacion"].apply(fmt_timedelta)

                    cols_esv = [c for c in [
                        "Antigüedad", "Jefe directo", "Vendedor", "Cliente", "Telefono", "Folio",
                        "Centro", "Estatus", "Venta", "Fecha creacion", "Back Office",
                    ] if c in df_esv.columns]

                    show_esv = df_esv[cols_esv].copy().rename(columns={"Vendedor": "Ejecutivo"})
                    show_esv = show_esv.assign(_age=df_esv["TD_Age_Desde_Creacion"]).sort_values("_age", ascending=False).drop(columns=["_age"], errors="ignore")

                    st.dataframe(show_esv, use_container_width=True, hide_index=True)

                    st.download_button(
                        "Descargar Entregado sin Venta (Excel)",
                        data=dfs_to_excel_bytes({"Entregado_sin_Venta": show_esv}),
                        file_name=f"entregado_sin_venta_{fecha_ini}_{fecha_fin}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

            st.caption(f"Actualizado: {st.session_state['last_refresh'].strftime('%Y-%m-%d %H:%M')}")

    # ============================
    # TAB 1: GRÁFICAS
    # ============================
    with tabs[1]:
        if df.empty:
            st.info("No hay datos para los filtros seleccionados.")
        else:
            view, meta = build_view(df, fecha_ini, fecha_fin)
            counts = compute_flow_counts(view)

            st.subheader("📊 Gráficas (enfoque jefe)")

            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(make_funnel(counts), use_container_width=True, key="t1_funnel")
            with c2:
                st.plotly_chart(make_flow_bar(counts), use_container_width=True, key="t1_flow_bar")

            fig_buckets = make_time_buckets_chart(view)
            if fig_buckets is not None:
                st.plotly_chart(fig_buckets, use_container_width=True, key="t1_buckets")

            fig_sc = make_scatter_orders(view, color_by=("Centro Original" if "Centro Original" in view.columns else None))
            if fig_sc is not None:
                st.plotly_chart(fig_sc, use_container_width=True, key="t1_scatter")

            with st.expander("Extras (solo si lo necesitas)", expanded=False):
                fig_backlog = make_backlog_over_time(view)
                if fig_backlog is not None:
                    st.plotly_chart(fig_backlog, use_container_width=True, key="t1_backlog")

                fig_hm = make_heatmap_created(view)
                if fig_hm is not None:
                    st.plotly_chart(fig_hm, use_container_width=True, key="t1_heatmap")

    # ============================
    # TAB 2: PENDIENTES A RECUPERAR
    # ============================
    with tabs[2]:
        if df.empty:
            st.info("No hay datos para los filtros seleccionados.")
        else:
            view, _ = build_view(df, fecha_ini, fecha_fin)
            work = view.copy()
            E = work["Estatus"].astype(str)

            work["Aging_h"] = np.where(
                E.eq("Back Office"),
                work["H_Age_Desde_BO"],
                work["H_Age_Desde_Creacion"],
            )

            work["CRITICO"] = False
            for stg, th in alert_map.items():
                work.loc[E.eq(stg) & (work["Aging_h"] > float(th)), "CRITICO"] = True

            crit = work[work["CRITICO"]].copy().sort_values("Aging_h", ascending=False)

            st.subheader("📌 Pendientes críticos (para recuperar hoy)")
            st.caption("Ordenados por antigüedad (más viejos arriba).")
            st.write(f"Total críticos: **{len(crit)}**")

            cols = [c for c in [
                "Estatus",
                "Jefe directo", "Vendedor", "Cliente", "Telefono", "Folio", "Centro", "Venta",
                "TD_Age_Desde_Creacion", "TD_Age_Desde_BO",
            ] if c in crit.columns]

            show = crit[cols].copy().rename(columns={"Vendedor": "Ejecutivo"})
            show["Antigüedad"] = np.where(
                show["Estatus"].astype(str).eq("Back Office"),
                crit["TD_Age_Desde_BO"].apply(fmt_timedelta),
                crit["TD_Age_Desde_Creacion"].apply(fmt_timedelta),
            )
            show.drop(columns=["TD_Age_Desde_Creacion", "TD_Age_Desde_BO"], inplace=True, errors="ignore")

            st.dataframe(show, use_container_width=True, hide_index=True)

            if not crit.empty:
                by_stage = crit.groupby("Estatus", as_index=False).size().rename(columns={"size": "Críticos"})
                by_stage["Estatus"] = pd.Categorical(by_stage["Estatus"], categories=FLOW_STAGES_NO_TOTAL, ordered=True)
                by_stage = by_stage.sort_values("Estatus")
                
                fig = px.bar(by_stage, x="Estatus", y="Críticos", title="Críticos por etapa", text_auto=True, template="plotly_white")
                fig.update_xaxes(type="category")
                st.plotly_chart(fig, use_container_width=True, key="t2_criticos_por_etapa")

            st.download_button(
                "Descargar críticos (Excel)",
                data=dfs_to_excel_bytes({"Criticos": show}),
                file_name=f"pendientes_criticos_{fecha_ini}_{fecha_fin}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ============================
    # TAB 3: DETALLE / EXPORT
    # ============================
    with tabs[3]:
        if df.empty:
            st.info("No hay datos para los filtros seleccionados.")
        else:
            view, meta = build_view(df, fecha_ini, fecha_fin)
            detail = view.copy()

            detail["Tiempo Nuevo→BO (HH:MM)"] = [
                (fmt_timedelta(done) if (done is not None and pd.notna(done))
                 else (f"En proceso · {fmt_timedelta(age)}" if (str(stg) == "Nuevo" and age is not None and pd.notna(age)) else "—"))
                for done, age, stg in zip(detail["TD_Creacion_a_BO"], detail["TD_Age_Desde_Creacion"], detail["Estatus"])
            ]
            detail["Tiempo BO→Entregado (HH:MM)"] = [
                fmt_done_or_in_process(done, age)
                for done, age in zip(detail["TD_BO_a_Entregado"], detail["TD_Age_Desde_BO"])
            ]
            detail["Tiempo Total a Entregado (HH:MM)"] = [
                fmt_done_or_in_process(done, age)
                for done, age in zip(detail["TD_Creacion_a_Entregado"], detail["TD_Age_Desde_Creacion"])
            ]

            detail["Tiempo BO→Act (HH:MM)"] = [
                fmt_done_or_in_process(done, age)
                for done, age in zip(detail["TD_BO_a_Act"], detail["TD_Age_Desde_BO"])
            ]
            detail["Tiempo Total (HH:MM)"] = [
                fmt_done_or_in_process(done, age)
                for done, age in zip(detail["TD_Creacion_a_Act"], detail["TD_Age_Desde_Creacion"])
            ]

            detail["Antigüedad desde Creación (HH:MM)"] = detail["TD_Age_Desde_Creacion"].apply(fmt_timedelta)
            detail["Antigüedad desde BO (HH:MM)"] = detail["TD_Age_Desde_BO"].apply(fmt_timedelta)

            for c in ["STG_Solicitado_DT", "STG_EnPreparacion_DT", "STG_EnEntrega_DT", "STG_Reprogramado_DT", "STG_Entregado_DT"]:
                if c in detail.columns:
                    detail[c] = pd.to_datetime(detail[c], errors="coerce")
                    detail[c] = sanitize_dates(detail[c])

            keep = [c for c in [
                "Estatus", "Jefe directo", "Vendedor", "Cliente", "Telefono", "Folio", "Centro",
                "Venta", "CREATED_DT", "BO_DT", "STG_Entregado_DT", "ACT_DT",
                "Tiempo Nuevo→BO (HH:MM)",
                "Tiempo BO→Entregado (HH:MM)",
                "Tiempo Total a Entregado (HH:MM)",
                "Tiempo BO→Act (HH:MM)",
                "Tiempo Total (HH:MM)",
                "Antigüedad desde Creación (HH:MM)", "Antigüedad desde BO (HH:MM)",
                "ENTREGADO_SIN_VENTA",
            ] if c in detail.columns]

            rename_map = {
                "Vendedor": "Ejecutivo",
                "CREATED_DT": "Fecha Creacion",
                "BO_DT": "Fecha Back Office",
                "STG_Entregado_DT": "Fecha Entregado",
                "ACT_DT": "Fecha Activacion",
            }

            show = detail[keep].copy().rename(columns=rename_map)
            if "Fecha Creacion" in show.columns:
                show = show.sort_values("Fecha Creacion", ascending=False)

            st.subheader("📄 Detalle completo")
            st.dataframe(show, use_container_width=True, hide_index=True)

            summary = pd.DataFrame([{
                "Periodo": f"{fecha_ini} a {fecha_fin}",
                "Órdenes total": int(len(view)),
                "Fuente Back Office": "Back Office (Rastreo)",
                "Fuente fecha activación": meta["activation_col"] if meta["has_activation_dt"] else "No disponible",
                "Actualizado": st.session_state["last_refresh"].strftime("%Y-%m-%d %H:%M"),
            }])

            st.download_button(
                "Descargar reporte (Excel)",
                data=dfs_to_excel_bytes({"Resumen": summary, "Detalle": show}),
                file_name=f"reporte_flujo_tiempos_{fecha_ini}_{fecha_fin}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
